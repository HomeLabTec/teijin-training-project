from flask import Flask, render_template, request, jsonify, redirect, url_for, session
import openpyxl
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "training-master.xlsx")

# list of stations for the schedule page
STATIONS = [
    "Press 1", "Press 2", "Press 3", "Press 4", "Press 5", "Press 6",
    "Press 7", "Press 8", "Press 9", "Press 10", "Press 11", "Press 12",
    "Press 13", "Press 20", "Press 25", "Press 26",
    "JT Bonder 1", "JT Bonder 2",
    "Water Jet 2", "Water Jet 3",
    "Great White Bonder 1", "Great White Bonder 2", "Great White Sub Bonder",
    "H567 Packout", "D-shield", "H567 Hood Bonder", "CS1",
    "H567 Ext Bonder", "Driver Assignment", "PC Assignment", "Common Load",
]

# map schedule stations to training columns in the spreadsheet
STATION_TO_HEADER = {
    "JT Bonder 1": "JT Bonder",
    "JT Bonder 2": "JT Bonder",
    "Great White Bonder 1": "GW Bonder",
    "Great White Bonder 2": "GW Bonder",
    "Great White Sub Bonder": "GW Bonder",
    "H567 Hood Bonder": "H567 Bonder",
    "H567 Ext Bonder": "H567 Bonder",
}

def load_workbook_data():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # detect skill columns (every other starting at C = 3)
    headers = []
    for skill_col in range(3, ws.max_column + 1, 2):
        meta_col = skill_col - 1
        # pull part name from row 1, part number from row 2
        part_name   = ws.cell(row=1, column=meta_col).value or ""
        part_number = ws.cell(row=2, column=meta_col).value or ""
        # build a display string: "Name (Number)" or fallback to whichever
        if part_name and part_number:
            display = f"{part_name} ({part_number})"
        else:
            display = part_name or str(part_number)
        headers.append((skill_col, display))

    # build data dict: { name → { display → level, … } }
    data = {}
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        skills = {}
        for col, display in headers:
            raw = ws.cell(row=row, column=col).value
            skills[display] = int(raw) if (raw is not None) else 1
        data[name] = skills

    return wb, ws, headers, data


def build_level_lookup(data):
    """Return mapping of name -> {station -> level} with default 1."""
    levels = {}
    for name, skills in data.items():
        per_station = {}
        for st in STATIONS:
            header = STATION_TO_HEADER.get(st)
            if header:
                per_station[st] = skills.get(header, 1)
            else:
                per_station[st] = 1
        levels[name] = per_station
    return levels


app = Flask(__name__)
app.secret_key = "dev"

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/display")
def display():
    # load headers & data as before
    wb, ws, headers, data = load_workbook_data()
    parts = [part for _, part in headers]
    return render_template("display.html", parts=parts, data=data)


@app.route("/search")
def search():
    q = request.args.get("q", "").lower()
    _, _, _, data = load_workbook_data()
    matches = [n for n in data if q in n.lower()]
    return jsonify(matches)

@app.route("/person/<name>", methods=["GET", "POST"])
def person(name):
    wb, ws, headers, data = load_workbook_data()

    if request.method == "POST":
        # ---- Handle Delete ----
        if request.form.get("action") == "delete":
            # find the row for this person and delete it
            for row in range(3, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == name:
                    ws.delete_rows(row, 1)
                    wb.save(EXCEL_PATH)
                    break
            return redirect(url_for("index"))

        # ---- Otherwise Handle Update ----
        for col, part in headers:
            new_val = int(request.form.get(part, 1))
            for row in range(3, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == name:
                    ws.cell(row=row, column=col, value=new_val)
                    break
        wb.save(EXCEL_PATH)
        return redirect(url_for("person", name=name))

    skills = data.get(name)
    if skills is None:
        return "Not found", 404

    parts = [p for _, p in headers]
    return render_template("person.html", name=name, parts=parts, skills=skills)


@app.route("/add", methods=["GET", "POST"])
def add():
    wb, ws, headers, _ = load_workbook_data()
    parts = [p for _, p in headers]

    if request.method == "POST":
        name = request.form["name"].strip()
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=name)
        for col, part in headers:
            ws.cell(row=new_row, column=col, value=int(request.form.get(part, 1)))
        wb.save(EXCEL_PATH)
        return redirect(url_for("index"))

    return render_template("add.html", parts=parts)


@app.route("/schedule")
def schedule():
    """Display a table of stations with a dropdown of workers for each."""
    _, _, _, data = load_workbook_data()
    names = sorted(data.keys())
    levels = build_level_lookup(data)
    stations = list(enumerate(STATIONS))
    return render_template("schedule.html", stations=stations, names=names, levels=levels)


@app.route("/generate_schedule", methods=["POST"])
def generate_schedule():
    """Store the submitted schedule and redirect to view page."""
    schedule = {}
    for idx, station in enumerate(STATIONS):
        people = []
        for j in range(6):
            key = f"station{idx}_{j}"
            people.append(request.form.get(key, ""))
        schedule[station] = people
    session['last_schedule'] = schedule
    return redirect(url_for('view_schedule'))


@app.route("/view_schedule")
def view_schedule():
    schedule = session.get('last_schedule')
    if not schedule:
        return redirect(url_for('schedule'))
    _, _, _, data = load_workbook_data()
    levels = build_level_lookup(data)
    return render_template("generated_schedule.html", schedule=schedule, levels=levels)

@app.route("/decrease", methods=["GET", "POST"])
def decrease():
    wb, ws, headers, _ = load_workbook_data()
    parts = [p for _, p in headers]

    if request.method == "POST":
        part   = request.form["part"]
        amount = int(request.form["amount"])
        col = next(c for c, p in headers if p == part)
        for row in range(3, ws.max_row + 1):
            cur = ws.cell(row=row, column=col).value or 1
            ws.cell(row=row, column=col, value=max(1, cur - amount))
        wb.save(EXCEL_PATH)
        return redirect(url_for("index"))

    return render_template("decrease.html", parts=parts)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
